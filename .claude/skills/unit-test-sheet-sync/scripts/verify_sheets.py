#!/usr/bin/env python3
"""Gate every write to Report5.1_Unit Test.xlsx: run this and fix until it prints 0 problems.

    py verify_sheets.py <file.xlsx> [sheet ...]      # no sheet list = every method sheet

Checks, in the order they usually break:
  form      three block labels in column A, one sub-header per parameter in Input,
            each case ticks at most one value per parameter group
  coverage  every case has a precondition and an expected outcome; no tick on an empty row;
            no row with text but no tick
  grid      the last case column is ruled on EVERY row of the table, not just the header row
  tail      nothing below "Defect ID" still carries table styling (the leftover blue band)
  date      "Executed Date" shows a date, not a raw serial number
  language  no Vietnamese outside quoted code messages
  totals    N + A + B == Total == number of case columns

Exit code is the number of problems, so it can gate a publish step.
"""
import io
import re
import sys
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SKIP = {"Cover", "MethodList", "Statistics", "methodName1", "Guideline"}
OUTSIDE_STYLES = {"35", "95", None}
SCAN_TO = 130
TYPE_PREFIX = "Type("
VN = re.compile(r"[ăâđêôơưĂÂĐÊÔƠƯáàảãạấầẩẫậắằẳẵặéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵ]")


def sheet_paths(z):
    wbxml = z.read("xl/workbook.xml").decode("utf-8")
    rels = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"',
                           z.read("xl/_rels/workbook.xml.rels").decode("utf-8")))
    out = {}
    for tag in re.findall(r"<sheet [^>]*/>", wbxml):
        name = re.search(r'name="([^"]+)"', tag).group(1)
        out[name] = "xl/" + rels[re.search(r'r:id="([^"]+)"', tag).group(1)].lstrip("/")
    return out


def check(name, ws, xml):
    rows_xml = {int(m.group(1)): m.group(2)
                for m in re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', xml, re.S)}

    def style(row, col):
        c = re.search(r'<c r="%s%d" s="(\d+)"' % (col, row), rows_xml.get(row, ""))
        return c.group(1) if c else None

    problems = []
    labels_a, labels_b = {}, {}
    for row in range(1, SCAN_TO):
        a, b = ws.cell(row, 1).value, ws.cell(row, 2).value
        if a:
            labels_a.setdefault(str(a).strip(), row)
        if b:
            labels_b.setdefault(str(b).strip(), row)

    for need in ("Condition", "Input", "Confirm", "Result"):
        if need not in labels_a:
            problems.append("missing block label in column A: %s" % need)
    defect = labels_b.get("Defect ID")
    date_row = labels_b.get("Executed Date")
    type_row = next((r for lb, r in labels_b.items() if lb.startswith(TYPE_PREFIX)), None)
    if not (defect and type_row):
        problems.append("missing Result rows")
        return problems

    cases = [c for c in range(6, 45) if ws.cell(7, c).value]
    in_start = labels_a.get("Input", type_row)
    conf = labels_a.get("Confirm", type_row)

    groups = []
    for row in range(in_start, conf):
        if ws.cell(row, 2).value:
            groups.append([str(ws.cell(row, 2).value).strip(), row, None])
    for i, g in enumerate(groups):
        g[2] = groups[i + 1][1] if i + 1 < len(groups) else conf
    if not groups:
        problems.append("Input block has no parameter sub-header")

    for col in cases:
        code = ws.cell(7, col).value
        for gname, start, end in groups:
            ticks = [r for r in range(start, end) if ws.cell(r, col).value and ws.cell(r, 4).value]
            if len(ticks) > 1:
                problems.append("%s picks %d values in group '%s'" % (code, len(ticks), gname))
        if not any(ws.cell(r, col).value and ws.cell(r, 4).value for r in range(9, in_start)):
            problems.append("%s has no precondition" % code)
        if not any(ws.cell(r, col).value and ws.cell(r, 4).value for r in range(conf, type_row)):
            problems.append("%s has no expected outcome" % code)

    for row in range(9, type_row):
        text = ws.cell(row, 4).value
        ticks = [c for c in range(6, 45) if ws.cell(row, c).value]
        if ticks and not text and not ws.cell(row, 2).value and not ws.cell(row, 1).value:
            problems.append("tick on empty row %d" % row)
        if text and not ticks:
            problems.append("row %d has text but no tick" % row)
        if text and VN.search(re.sub(r'"[^"]*"', "", str(text))):
            problems.append("Vietnamese outside quotes on row %d" % row)

    if cases:
        last = get_column_letter(cases[-1])
        off = [r for r in range(7, defect + 1)
               if r in rows_xml and style(r, "F") is not None and style(r, last) != style(r, "F")]
        if off:
            problems.append("last case column %s not ruled on rows %s" % (last, off[:10]))

    leftovers = [r for r in range(defect + 1, SCAN_TO)
                 if style(r, "A") not in OUTSIDE_STYLES or style(r, "B") not in OUTSIDE_STYLES]
    if leftovers:
        problems.append("table styling still below Defect ID on rows %s" % leftovers[:10])

    if date_row and isinstance(ws.cell(date_row, 6).value, (int, float)):
        problems.append("Executed Date shows a serial number, not a date")

    n, a, b, total = (ws.cell(5, c).value for c in (12, 13, 14, 15))
    if not (isinstance(total, int) and n + a + b == total == len(cases)):
        problems.append("header counts %s+%s+%s vs total %s vs %d case columns"
                        % (n, a, b, total, len(cases)))
    return problems


def main():
    path = sys.argv[1]
    wanted = sys.argv[2:]
    z = zipfile.ZipFile(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    paths = sheet_paths(z)

    print("entries: %d | sheets: %d" % (len(z.namelist()), len(wb.sheetnames)))
    total = 0
    for name in wb.sheetnames:
        if name in SKIP or (wanted and name not in wanted):
            continue
        problems = check(name, wb[name], z.read(paths[name]).decode("utf-8"))
        total += len(problems)
        print("%-32s %s" % (name, "OK" if not problems else "; ".join(problems)))
    print("TOTAL PROBLEMS:", total)
    sys.exit(min(total, 120))


if __name__ == "__main__":
    main()
